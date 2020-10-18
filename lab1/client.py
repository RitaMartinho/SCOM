#!/usr/bin/python

import os, xlsxwriter, subprocess, time
from openpyxl import workbook #pip install openpyxl
from openpyxl import load_workbook
from datetime import datetime

number= input('Choose a N for N calls to the service:')
i=0

wb = load_workbook(filename = 'lab1.xlsx')
sheet= wb['Sheet1']

first_column = sheet['A']

if len(first_column)==0:
    starting_row=1
else:
    starting_row=len(first_column)+1

while(i<number):

    os.system("sleep 3")
    now =datetime.now()
    
    cmd= 'curl -o /dev/null -s -w "%{http_code}" --data "value=100" http://35.228.185.222/cgi-bin/lab1.py'
    http_code=subprocess.check_output(cmd, shell=True)

    if(http_code == "200"):
       
        after = datetime.now()

        sheet.cell(row=starting_row, column= 1).value= (after-now).total_seconds()*1000 #write in miliseconfs

        starting_row += 1
    i += 1

sheet.cell(row=1, column=3).value = "Average: "
sheet.cell(row=1, column=4).value = '=AVERAGE(A:A)'
sheet.cell(row=1, column=5).value = "Std Deviation: "
sheet.cell(row=1, column=6).value = "=STDEV(A:A)"
sheet.cell(row=1, column=7).value = "Confidence:"
sheet.cell(row=1, column=8).value = "=CONFIDENCE(0.05,STDEV(A:A), COUNT(A:A))"
sheet.cell(row=3, column=3).value = "5% of average:"
sheet.cell(row=3, column=4).value = "=0.05*D1"



wb.save('lab1.xlsx')
